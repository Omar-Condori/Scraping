from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import time
import re
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def obtener_fechas_semana():
    """Obtiene las fechas de inicio y fin de la semana actual"""
    hoy = datetime.now()
    inicio_semana = hoy - timedelta(days=hoy.weekday())  # Lunes
    fin_semana = inicio_semana + timedelta(days=6)  # Domingo
    
    return inicio_semana, fin_semana

def extraer_categoria_desde_url(url):
    """Extrae la categoría basándose en la URL de la noticia"""
    
    categorias_url = {
        '/politica/': 'Política',
        '/economia/': 'Economía', 
        '/mundo/': 'Mundo',
        '/deportes/': 'Deportes',
        '/espectaculos/': 'Espectáculos',
        '/tecnologia/': 'Tecnología',
        '/sociedad/': 'Sociedad',
        '/cultura/': 'Cultura',
        '/salud/': 'Salud',
        '/educacion/': 'Educación',
        '/turismo/': 'Turismo',
        '/gastronomia/': 'Gastronomía',
        '/saltar-intro/': 'Entretenimiento',
        '/noticias/': 'Noticias Generales'
    }
    
    for path, categoria in categorias_url.items():
        if path in url:
            return categoria
    
    return 'Sin categoría'

def extraer_categoria_desde_titulo(titulo):
    """Extrae categoría basándose en palabras clave del título"""
    
    palabras_clave = {
        'congreso': 'Política',
        'presidente': 'Política', 
        'gobierno': 'Política',
        'ministro': 'Política',
        'elecciones': 'Política',
        'partido': 'Política',
        'dólar': 'Economía',
        'inflación': 'Economía',
        'empresa': 'Economía',
        'mercado': 'Economía',
        'banco': 'Economía',
        'futbol': 'Deportes',
        'fútbol': 'Deportes',
        'deporte': 'Deportes',
        'mundial': 'Deportes',
        'liga': 'Deportes',
        'película': 'Espectáculos',
        'serie': 'Espectáculos',
        'actor': 'Espectáculos',
        'música': 'Espectáculos',
        'anime': 'Entretenimiento',
        'manga': 'Entretenimiento',
        'temporada': 'Entretenimiento',
        'capítulo': 'Entretenimiento',
        'covid': 'Salud',
        'salud': 'Salud',
        'hospital': 'Salud',
        'vacuna': 'Salud',
        'terremoto': 'Sociedad',
        'accidente': 'Sociedad',
        'crimen': 'Sociedad',
        'robo': 'Sociedad'
    }
    
    titulo_lower = titulo.lower()
    for palabra, categoria in palabras_clave.items():
        if palabra in titulo_lower:
            return categoria
    
    return None

def extraer_contenido_noticia_mejorado(driver, url):
    """Extrae el contenido completo de una noticia individual con múltiples estrategias"""
    try:
        print(f"    📖 Extrayendo contenido de: {url[:60]}...")
        driver.get(url)
        
        # Esperar que la página cargue completamente
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        time.sleep(3)
        
        soup = BeautifulSoup(driver.page_source, "html.parser")
        
        # Estrategia 1: Buscar selectores específicos de El Comercio
        selectores_contenido = [
            ".story-content",
            ".article-content", 
            ".noticia-content",
            ".story-body",
            ".article-body",
            ".content",
            ".texto",
            ".parrafo",
            "article .text",
            ".story-text",
            ".article-text",
            ".entry-content",
            ".post-content",
            ".news-content",
            ".main-content",
            ".article-main",
            ".story-main"
        ]
        
        contenido_completo = ""
        resumen = ""
        
        # Buscar resumen o lead
        selectores_resumen = [
            ".story-summary",
            ".article-summary",
            ".lead",
            ".resumen",
            ".abstract",
            ".story-lead",
            ".article-lead",
            ".excerpt",
            ".summary"
        ]
        
        for selector in selectores_resumen:
            resumen_elem = soup.select_one(selector)
            if resumen_elem:
                resumen = resumen_elem.get_text(strip=True)
                break
        
        # Estrategia 2: Buscar contenido principal
        for selector in selectores_contenido:
            contenido_elem = soup.select_one(selector)
            if contenido_elem:
                # Extraer párrafos
                parrafos = contenido_elem.find_all(['p', 'div'], string=True)
                for parrafo in parrafos:
                    texto = parrafo.get_text(strip=True)
                    if len(texto) > 20:  # Filtrar párrafos muy cortos
                        contenido_completo += texto + "\n\n"
                break
        
        # Estrategia 3: Si no se encontró contenido específico, buscar en todo el artículo
        if not contenido_completo:
            article = soup.find('article')
            if article:
                parrafos = article.find_all(['p', 'div'], string=True)
                for parrafo in parrafos:
                    texto = parrafo.get_text(strip=True)
                    if len(texto) > 20:
                        contenido_completo += texto + "\n\n"
        
        # Estrategia 4: Buscar en el main content
        if not contenido_completo:
            main = soup.find('main')
            if main:
                parrafos = main.find_all(['p', 'div'], string=True)
                for parrafo in parrafos:
                    texto = parrafo.get_text(strip=True)
                    if len(texto) > 20:
                        contenido_completo += texto + "\n\n"
        
        # Estrategia 5: Buscar todos los párrafos en la página
        if not contenido_completo:
            parrafos = soup.find_all('p')
            for parrafo in parrafos:
                texto = parrafo.get_text(strip=True)
                if len(texto) > 50:  # Párrafos más largos
                    contenido_completo += texto + "\n\n"
        
        # Estrategia 6: Usar JavaScript para obtener texto
        if not contenido_completo:
            try:
                # Ejecutar JavaScript para obtener texto de la página
                js_text = driver.execute_script("""
                    var text = '';
                    var elements = document.querySelectorAll('p, div, article, main');
                    for (var i = 0; i < elements.length; i++) {
                        var elementText = elements[i].innerText || elements[i].textContent;
                        if (elementText && elementText.length > 50) {
                            text += elementText + '\\n\\n';
                        }
                    }
                    return text;
                """)
                if js_text and len(js_text) > 100:
                    contenido_completo = js_text
            except:
                pass
        
        # Limpiar contenido
        contenido_completo = re.sub(r'\s+', ' ', contenido_completo).strip()
        resumen = re.sub(r'\s+', ' ', resumen).strip()
        
        # Filtrar contenido no relevante
        contenido_completo = re.sub(r'(Publicidad|Anuncio|Sponsor|Cookie|Política de privacidad).*', '', contenido_completo, flags=re.IGNORECASE)
        
        # Limitar longitud
        if len(contenido_completo) > 8000:
            contenido_completo = contenido_completo[:8000] + "..."
        
        # Si no hay resumen, usar las primeras líneas del contenido
        if not resumen and contenido_completo:
            resumen = contenido_completo[:200] + "..." if len(contenido_completo) > 200 else contenido_completo
        
        return {
            'resumen': resumen,
            'contenido': contenido_completo,
            'longitud': len(contenido_completo)
        }
        
    except Exception as e:
        print(f"    ❌ Error extrayendo contenido: {e}")
        return {
            'resumen': "Error al extraer resumen",
            'contenido': "Error al extraer contenido",
            'longitud': 0
        }

# Configuración de Chrome
chrome_options = Options()
chrome_options.add_argument("--headless=new")  
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument("--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

service = Service("/opt/homebrew/bin/chromedriver")
driver = webdriver.Chrome(service=service, options=chrome_options)

try:
    # Obtener fechas de la semana
    inicio_semana, fin_semana = obtener_fechas_semana()
    print(f"📅 Buscando noticias de la semana: {inicio_semana.strftime('%d/%m/%Y')} - {fin_semana.strftime('%d/%m/%Y')}")
    
    # Abrir página de últimas noticias
    print("🔍 Accediendo a El Comercio...")
    driver.get("https://elcomercio.pe/ultimas-noticias/")
    time.sleep(8)
    
    # Obtener el HTML de la página
    soup = BeautifulSoup(driver.page_source, "html.parser")
    
    print("📰 Analizando estructura de la página...")
    
    # Buscar diferentes tipos de selectores
    selectores_posibles = [
        "article",
        ".story-item",
        ".story-card", 
        ".Card-link",
        ".StoryLink",
        ".stories__item-link",
        "a[href*='/noticias/']",
        "a[href*='/politica/']",
        "a[href*='/economia/']",
        "a[href*='/mundo/']",
        "a[href*='/deportes/']",
        "a[href*='/espectaculos/']",
        ".headline",
        ".title",
        "h1 a",
        "h2 a", 
        "h3 a",
        ".news-item",
        ".noticia"
    ]
    
    noticias_encontradas = []
    
    for selector in selectores_posibles:
        elementos = soup.select(selector)
        if elementos:
            print(f"✅ Selector '{selector}' encontró {len(elementos)} elementos")
            
            for elem in elementos[:8]:  # Limitar para no sobrecargar
                # Buscar enlaces y títulos
                if elem.name == 'a':
                    enlace = elem.get('href', '')
                    titulo = elem.get_text(strip=True)
                else:
                    enlace_elem = elem.find('a')
                    if enlace_elem:
                        enlace = enlace_elem.get('href', '')
                        titulo = enlace_elem.get_text(strip=True)
                    else:
                        continue
                
                # Filtrar enlaces válidos de noticias
                if (enlace and 
                    ('/noticias/' in enlace or '/politica/' in enlace or 
                     '/economia/' in enlace or '/mundo/' in enlace or
                     '/deportes/' in enlace or '/espectaculos/' in enlace or
                     '/saltar-intro/' in enlace) and
                    titulo and len(titulo) > 10 and
                    not any(palabra in titulo.lower() for palabra in ['publicidad', 'anuncio', 'sponsor', 'cookies'])):
                    
                    # Completar URL si es relativa
                    if enlace.startswith("/"):
                        enlace = "https://elcomercio.pe" + enlace
                    
                    # Determinar categoría
                    categoria_url = extraer_categoria_desde_url(enlace)
                    categoria_titulo = extraer_categoria_desde_titulo(titulo)
                    categoria_final = categoria_titulo if categoria_titulo else categoria_url
                    
                    noticia = {
                        'titulo': titulo,
                        'enlace': enlace,
                        'categoria': categoria_final,
                        'fecha_extraccion': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    }
                    
                    noticias_encontradas.append(noticia)
    
    # Eliminar duplicados
    noticias_unicas = []
    titulos_vistos = set()
    
    for noticia in noticias_encontradas:
        titulo_normalizado = re.sub(r'[^\w\s]', '', noticia['titulo'].lower())
        if titulo_normalizado not in titulos_vistos:
            titulos_vistos.add(titulo_normalizado)
            noticias_unicas.append(noticia)
    
    # Limitar a las 8 noticias más importantes para extraer contenido
    noticias_para_extraer = noticias_unicas[:8]
    
    print(f"\n📖 Extrayendo contenido de {len(noticias_para_extraer)} noticias...")
    
    # Extraer contenido de cada noticia
    noticias_con_contenido = []
    for i, noticia in enumerate(noticias_para_extraer, 1):
        print(f"\n📰 Noticia {i}/{len(noticias_para_extraer)}: {noticia['titulo'][:50]}...")
        
        contenido_info = extraer_contenido_noticia_mejorado(driver, noticia['enlace'])
        
        noticia_completa = {
            **noticia,
            'resumen': contenido_info['resumen'],
            'contenido': contenido_info['contenido'],
            'longitud_contenido': contenido_info['longitud']
        }
        
        noticias_con_contenido.append(noticia_completa)
        
        # Pausa entre extracciones
        time.sleep(2)
    
    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Noticias con Contenido"
    
    # Agregar encabezado
    ws['A1'] = "REPORTE DE NOTICIAS CON CONTENIDO COMPLETO - EL COMERCIO"
    ws['A2'] = f"Semana: {inicio_semana.strftime('%d/%m/%Y')} - {fin_semana.strftime('%d/%m/%Y')}"
    ws['A3'] = f"Fecha de extracción: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A4'] = f"Total de noticias con contenido: {len(noticias_con_contenido)}"
    
    # Estilos para el encabezado
    header_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Agregar datos desde la fila 6
    headers = ['#', 'TÍTULO', 'CATEGORÍA', 'RESUMEN', 'CONTENIDO COMPLETO', 'LONGITUD', 'ENLACE']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Agregar datos de noticias
    for idx, noticia in enumerate(noticias_con_contenido, 1):
        row = idx + 6
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=noticia['titulo'])
        ws.cell(row=row, column=3, value=noticia['categoria'])
        ws.cell(row=row, column=4, value=noticia['resumen'])
        ws.cell(row=row, column=5, value=noticia['contenido'])
        ws.cell(row=row, column=6, value=noticia['longitud_contenido'])
        ws.cell(row=row, column=7, value=noticia['enlace'])
        
        # Alternar colores de fila
        if idx % 2 == 0:
            fill_color = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = fill_color
    
    # Ajustar ancho de columnas
    column_widths = [5, 60, 15, 50, 100, 10, 60]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Ajustar altura de filas para contenido
    for row in range(7, len(noticias_con_contenido) + 7):
        ws.row_dimensions[row].height = 300  # Altura mayor para mejor visualización
    
    # Agregar bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(6, len(noticias_con_contenido) + 7):
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = thin_border
    
    # Crear hoja de resumen
    ws2 = wb.create_sheet("Resumen")
    
    # Contar noticias por categoría
    categorias_count = {}
    for noticia in noticias_con_contenido:
        cat = noticia['categoria']
        categorias_count[cat] = categorias_count.get(cat, 0) + 1
    
    # Agregar encabezado de resumen
    ws2['A1'] = "RESUMEN DE NOTICIAS CON CONTENIDO"
    ws2['A1'].font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    ws2['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws2['A1'].alignment = Alignment(horizontal='center')
    
    ws2['A3'] = "CATEGORÍA"
    ws2['B3'] = "CANTIDAD"
    ws2['C3'] = "PROMEDIO LONGITUD"
    ws2['A3'].font = Font(name='Arial', size=12, bold=True)
    ws2['B3'].font = Font(name='Arial', size=12, bold=True)
    ws2['C3'].font = Font(name='Arial', size=12, bold=True)
    
    # Agregar datos de categorías con estadísticas
    for idx, (categoria, count) in enumerate(sorted(categorias_count.items(), key=lambda x: x[1], reverse=True), 1):
        row = idx + 3
        
        # Calcular promedio de longitud por categoría
        noticias_categoria = [n for n in noticias_con_contenido if n['categoria'] == categoria]
        promedio_longitud = sum(n['longitud_contenido'] for n in noticias_categoria) // len(noticias_categoria)
        
        ws2.cell(row=row, column=1, value=categoria)
        ws2.cell(row=row, column=2, value=count)
        ws2.cell(row=row, column=3, value=promedio_longitud)
        
        if idx % 2 == 0:
            fill_color = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            for col in range(1, 4):
                ws2.cell(row=row, column=col).fill = fill_color
    
    # Ajustar ancho de columnas en resumen
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 20
    
    # Guardar archivo Excel
    nombre_archivo = f"reporte_contenido_completo_elcomercio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(nombre_archivo)
    
    print(f"\n📊 REPORTE CON CONTENIDO COMPLETO GENERADO EN EXCEL")
    print("=" * 80)
    print(f"📁 Archivo guardado como: {nombre_archivo}")
    print(f"📅 Período: {inicio_semana.strftime('%d/%m/%Y')} - {fin_semana.strftime('%d/%m/%Y')}")
    print(f"📈 Total de noticias con contenido: {len(noticias_con_contenido)}")
    print(f"📋 Categorías encontradas: {len(categorias_count)}")
    print("=" * 80)
    
    print(f"\n🔥 NOTICIAS CON CONTENIDO EXTRAÍDO:")
    print("-" * 80)
    
    for i, noticia in enumerate(noticias_con_contenido, 1):
        print(f"\n{i}. {noticia['titulo']}")
        print(f"   🏷️  Categoría: {noticia['categoria']}")
        print(f"   📝 Resumen: {noticia['resumen'][:100]}...")
        print(f"   📊 Longitud: {noticia['longitud_contenido']} caracteres")
        print(f"   🔗 {noticia['enlace']}")
        print("-" * 60)
    
    print(f"\n📈 RESUMEN POR CATEGORÍAS:")
    print("-" * 40)
    for categoria, count in sorted(categorias_count.items(), key=lambda x: x[1], reverse=True):
        noticias_cat = [n for n in noticias_con_contenido if n['categoria'] == categoria]
        promedio = sum(n['longitud_contenido'] for n in noticias_cat) // len(noticias_cat)
        print(f"   {categoria}: {count} noticias (promedio: {promedio} caracteres)")
    
    print(f"\n✅ ¡Archivo Excel con contenido completo listo para descargar!")
    print(f"📂 Ubicación: /Users/omar/Documents/Scraping/{nombre_archivo}")

except Exception as e:
    print(f"❌ Error: {e}")
    
finally:
    driver.quit()
    print("\n✅ Proceso completado.")
