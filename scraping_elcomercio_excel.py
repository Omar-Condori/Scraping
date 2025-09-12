from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import time
import re
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def buscar_fecha_en_elemento(elemento):
    """Busca fechas en diferentes formatos dentro de un elemento HTML"""
    
    # Selectores comunes para fechas
    selectores_fecha = [
        "time",
        ".date",
        ".fecha", 
        ".timestamp",
        ".publish-date",
        ".publication-date",
        "[datetime]",
        ".story-date",
        ".news-date"
    ]
    
    # Buscar en el elemento actual
    for selector in selectores_fecha:
        fecha_elem = elemento.select_one(selector)
        if fecha_elem:
            # Intentar obtener fecha del atributo datetime
            if fecha_elem.has_attr('datetime'):
                return fecha_elem['datetime']
            # O del texto del elemento
            fecha_texto = fecha_elem.get_text(strip=True)
            if fecha_texto:
                return fecha_texto
    
    # Buscar en elementos padre
    padre = elemento.parent
    while padre and padre.name != 'body':
        for selector in selectores_fecha:
            fecha_elem = padre.select_one(selector)
            if fecha_elem:
                if fecha_elem.has_attr('datetime'):
                    return fecha_elem['datetime']
                fecha_texto = fecha_elem.get_text(strip=True)
                if fecha_texto:
                    return fecha_texto
        padre = padre.parent
    
    # Buscar en elementos hermanos
    if elemento.parent:
        for hermano in elemento.parent.find_all():
            if hermano != elemento:
                for selector in selectores_fecha:
                    fecha_elem = hermano.select_one(selector)
                    if fecha_elem:
                        if fecha_elem.has_attr('datetime'):
                            return fecha_elem['datetime']
                        fecha_texto = fecha_elem.get_text(strip=True)
                        if fecha_texto:
                            return fecha_texto
    
    # Buscar patrones de fecha en el texto
    texto_completo = elemento.get_text()
    patrones_fecha = [
        r'\d{1,2}/\d{1,2}/\d{4}',
        r'\d{4}-\d{2}-\d{2}',
        r'\d{1,2} de \w+ de \d{4}',
        r'hace \d+ \w+',
        r'\d{1,2}:\d{2}',
        r'hoy',
        r'ayer'
    ]
    
    for patron in patrones_fecha:
        match = re.search(patron, texto_completo, re.IGNORECASE)
        if match:
            return match.group()
    
    return "Fecha no encontrada"

def extraer_categoria_desde_url(url):
    """Extrae la categoría basándose en la URL de la noticia"""
    
    # Mapeo de URLs a categorías
    categorias_url = {
        '/politica/': 'Política',
        '/economia/': 'Economía', 
        '/mundo/': 'Mundo',
        '/deportes/': 'Deportes',
        '/espectaculos/': 'Espectáculos',
        '/tecnologia/': 'Tecnología',
        '/sociedad/': 'Sociedad',
        '/cultura/': 'Cultura',
        '/salud/': 'Salud',
        '/educacion/': 'Educación',
        '/turismo/': 'Turismo',
        '/gastronomia/': 'Gastronomía',
        '/saltar-intro/': 'Entretenimiento',
        '/noticias/': 'Noticias Generales'
    }
    
    for path, categoria in categorias_url.items():
        if path in url:
            return categoria
    
    return 'Sin categoría'

def extraer_categoria_desde_titulo(titulo):
    """Extrae categoría basándose en palabras clave del título"""
    
    palabras_clave = {
        'congreso': 'Política',
        'presidente': 'Política', 
        'gobierno': 'Política',
        'ministro': 'Política',
        'elecciones': 'Política',
        'partido': 'Política',
        'dólar': 'Economía',
        'inflación': 'Economía',
        'empresa': 'Economía',
        'mercado': 'Economía',
        'banco': 'Economía',
        'futbol': 'Deportes',
        'fútbol': 'Deportes',
        'deporte': 'Deportes',
        'mundial': 'Deportes',
        'liga': 'Deportes',
        'película': 'Espectáculos',
        'serie': 'Espectáculos',
        'actor': 'Espectáculos',
        'música': 'Espectáculos',
        'anime': 'Entretenimiento',
        'manga': 'Entretenimiento',
        'temporada': 'Entretenimiento',
        'capítulo': 'Entretenimiento',
        'covid': 'Salud',
        'salud': 'Salud',
        'hospital': 'Salud',
        'vacuna': 'Salud',
        'terremoto': 'Sociedad',
        'accidente': 'Sociedad',
        'crimen': 'Sociedad',
        'robo': 'Sociedad'
    }
    
    titulo_lower = titulo.lower()
    for palabra, categoria in palabras_clave.items():
        if palabra in titulo_lower:
            return categoria
    
    return None

# Configuración de Chrome
chrome_options = Options()
chrome_options.add_argument("--headless=new")  
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument("--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

service = Service("/opt/homebrew/bin/chromedriver")
driver = webdriver.Chrome(service=service, options=chrome_options)

try:
    # Abrir página de últimas noticias
    print("🔍 Accediendo a El Comercio...")
    driver.get("https://elcomercio.pe/ultimas-noticias/")
    time.sleep(8)  # Esperar más tiempo para que cargue todo el contenido
    
    # Obtener el HTML de la página
    soup = BeautifulSoup(driver.page_source, "html.parser")
    
    print("📰 Analizando estructura de la página...")
    
    # Buscar diferentes tipos de selectores que podrían contener noticias
    selectores_posibles = [
        "article",
        ".story-item",
        ".story-card", 
        ".Card-link",
        ".StoryLink",
        ".stories__item-link",
        "a[href*='/noticias/']",
        "a[href*='/politica/']",
        "a[href*='/economia/']",
        "a[href*='/mundo/']",
        "a[href*='/deportes/']",
        "a[href*='/espectaculos/']",
        ".headline",
        ".title",
        "h1 a",
        "h2 a", 
        "h3 a",
        ".news-item",
        ".noticia"
    ]
    
    noticias_encontradas = []
    
    for selector in selectores_posibles:
        elementos = soup.select(selector)
        if elementos:
            print(f"✅ Selector '{selector}' encontró {len(elementos)} elementos")
            
            for elem in elementos[:10]:  # Limitar a 10 por selector
                # Buscar enlaces y títulos
                if elem.name == 'a':
                    enlace = elem.get('href', '')
                    titulo = elem.get_text(strip=True)
                    # Buscar fecha en el elemento padre o hermanos
                    fecha = buscar_fecha_en_elemento(elem)
                else:
                    enlace_elem = elem.find('a')
                    if enlace_elem:
                        enlace = enlace_elem.get('href', '')
                        titulo = enlace_elem.get_text(strip=True)
                        # Buscar fecha en el elemento actual
                        fecha = buscar_fecha_en_elemento(elem)
                    else:
                        continue
                
                # Filtrar enlaces válidos de noticias
                if (enlace and 
                    ('/noticias/' in enlace or '/politica/' in enlace or 
                     '/economia/' in enlace or '/mundo/' in enlace or
                     '/deportes/' in enlace or '/espectaculos/' in enlace or
                     '/saltar-intro/' in enlace) and
                    titulo and len(titulo) > 10 and
                    not any(palabra in titulo.lower() for palabra in ['publicidad', 'anuncio', 'sponsor', 'cookies'])):
                    
                    # Completar URL si es relativa
                    if enlace.startswith("/"):
                        enlace = "https://elcomercio.pe" + enlace
                    
                    # Determinar categoría
                    categoria_url = extraer_categoria_desde_url(enlace)
                    categoria_titulo = extraer_categoria_desde_titulo(titulo)
                    
                    # Priorizar categoría del título si es más específica
                    categoria_final = categoria_titulo if categoria_titulo else categoria_url
                    
                    noticias_encontradas.append({
                        'titulo': titulo,
                        'enlace': enlace,
                        'categoria': categoria_final,
                        'fecha': fecha,
                        'selector': selector,
                        'fecha_extraccion': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    })
    
    # Eliminar duplicados basándose en el título
    noticias_unicas = []
    titulos_vistos = set()
    
    for noticia in noticias_encontradas:
        titulo_normalizado = re.sub(r'[^\w\s]', '', noticia['titulo'].lower())
        if titulo_normalizado not in titulos_vistos:
            titulos_vistos.add(titulo_normalizado)
            noticias_unicas.append(noticia)
    
    # Crear DataFrame
    df = pd.DataFrame(noticias_unicas)
    
    # Crear archivo Excel con formato
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte El Comercio"
    
    # Agregar encabezado
    ws['A1'] = "REPORTE DE NOTICIAS - EL COMERCIO"
    ws['A2'] = f"Fecha de extracción: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'] = f"Total de noticias encontradas: {len(noticias_unicas)}"
    
    # Estilos para el encabezado
    header_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Agregar datos desde la fila 5
    headers = ['#', 'TÍTULO', 'CATEGORÍA', 'FECHA', 'ENLACE', 'SELECTOR', 'FECHA EXTRACCIÓN']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Agregar datos de noticias
    for idx, noticia in enumerate(noticias_unicas, 1):
        row = idx + 5
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=noticia['titulo'])
        ws.cell(row=row, column=3, value=noticia['categoria'])
        ws.cell(row=row, column=4, value=noticia['fecha'])
        ws.cell(row=row, column=5, value=noticia['enlace'])
        ws.cell(row=row, column=6, value=noticia['selector'])
        ws.cell(row=row, column=7, value=noticia['fecha_extraccion'])
        
        # Alternar colores de fila
        if idx % 2 == 0:
            fill_color = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = fill_color
    
    # Ajustar ancho de columnas
    column_widths = [5, 80, 15, 20, 60, 20, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Agregar bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(5, len(noticias_unicas) + 6):
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = thin_border
    
    # Crear hoja de resumen por categorías
    ws2 = wb.create_sheet("Resumen por Categorías")
    
    # Contar noticias por categoría
    categorias_count = {}
    for noticia in noticias_unicas:
        cat = noticia['categoria']
        categorias_count[cat] = categorias_count.get(cat, 0) + 1
    
    # Agregar encabezado de resumen
    ws2['A1'] = "RESUMEN POR CATEGORÍAS"
    ws2['A1'].font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    ws2['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws2['A1'].alignment = Alignment(horizontal='center')
    
    ws2['A3'] = "CATEGORÍA"
    ws2['B3'] = "CANTIDAD"
    ws2['A3'].font = Font(name='Arial', size=12, bold=True)
    ws2['B3'].font = Font(name='Arial', size=12, bold=True)
    
    # Agregar datos de categorías
    for idx, (categoria, count) in enumerate(sorted(categorias_count.items(), key=lambda x: x[1], reverse=True), 1):
        row = idx + 3
        ws2.cell(row=row, column=1, value=categoria)
        ws2.cell(row=row, column=2, value=count)
        
        if idx % 2 == 0:
            fill_color = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            for col in range(1, 3):
                ws2.cell(row=row, column=col).fill = fill_color
    
    # Ajustar ancho de columnas en resumen
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 15
    
    # Guardar archivo Excel
    nombre_archivo = f"reporte_elcomercio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(nombre_archivo)
    
    print(f"\n📊 REPORTE COMPLETO GENERADO EN EXCEL")
    print("=" * 60)
    print(f"📁 Archivo guardado como: {nombre_archivo}")
    print(f"📈 Total de noticias: {len(noticias_unicas)}")
    print(f"📋 Categorías encontradas: {len(categorias_count)}")
    print("=" * 60)
    
    print(f"\n🔥 LAS 5 NOTICIAS MÁS IMPORTANTES:")
    print("-" * 60)
    
    for i, noticia in enumerate(noticias_unicas[:5], 1):
        print(f"\n{i}. {noticia['titulo']}")
        print(f"   🏷️  Categoría: {noticia['categoria']}")
        print(f"   📅 Fecha: {noticia['fecha']}")
        print(f"   🔗 {noticia['enlace']}")
        print("-" * 40)
    
    print(f"\n📈 RESUMEN POR CATEGORÍAS:")
    print("-" * 30)
    for categoria, count in sorted(categorias_count.items(), key=lambda x: x[1], reverse=True):
        print(f"   {categoria}: {count} noticias")
    
    print(f"\n✅ ¡Archivo Excel listo para descargar!")
    print(f"📂 Ubicación: /Users/omar/Documents/Scraping/{nombre_archivo}")

except Exception as e:
    print(f"❌ Error: {e}")
    
finally:
    driver.quit()
    print("\n✅ Proceso completado.")
